"""Download all registrations into registrations.xlsx.

    pip install requests openpyxl
    python export-to-excel.py

Set SUPABASE_URL and SUPABASE_SERVICE_KEY below or as environment variables.
Use the *service_role* key, not the anon key — the table is insert-only for
anon, so the public key cannot read rows back.

Keep the service_role key off the website and out of git. It bypasses RLS.
"""

import os
import sys

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://YOUR-PROJECT.supabase.co")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "YOUR-SERVICE-ROLE-KEY")

OUTPUT = "registrations.xlsx"

COLUMNS = [
    ("submitted_at", "Submitted"),
    ("team_name", "Team"),
    ("team_tag", "Tag"),
    ("game", "Game"),
    ("captain_name", "Captain"),
    ("captain_email", "Email"),
    ("captain_phone", "Phone"),
    ("discord_id", "Discord"),
    ("player_1", "Player 1"),
    ("player_2", "Player 2"),
    ("player_3", "Player 3"),
    ("player_4", "Player 4"),
    ("player_5", "Player 5"),
    ("substitute", "Sub"),
    ("notes", "Notes"),
    ("agreed_rules", "Agreed"),
]


def fetch_rows():
    """Page through the table so exports don't silently stop at 1000 rows."""
    rows, step, offset = [], 1000, 0
    while True:
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/registrations",
            headers={
                "apikey": SUPABASE_SERVICE_KEY,
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "Range": f"{offset}-{offset + step - 1}",
            },
            params={"select": "*", "order": "submitted_at.asc"},
            timeout=30,
        )
        res.raise_for_status()
        batch = res.json()
        rows.extend(batch)
        if len(batch) < step:
            return rows
        offset += step


def main():
    if "YOUR-PROJECT" in SUPABASE_URL or "YOUR-SERVICE-ROLE" in SUPABASE_SERVICE_KEY:
        sys.exit("Set SUPABASE_URL and SUPABASE_SERVICE_KEY first (see docstring).")

    rows = fetch_rows()

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"

    ws.append([label for _, label in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append([row.get(key) for key, _ in COLUMNS])

    # Freeze the header and size columns to their contents so the sheet is
    # readable the moment it opens.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, (key, label) in enumerate(COLUMNS, start=1):
        widest = max([len(label)] + [len(str(r.get(key) or "")) for r in rows])
        ws.column_dimensions[get_column_letter(i)].width = min(widest + 3, 45)

    wb.save(OUTPUT)
    print(f"Wrote {len(rows)} registrations to {OUTPUT}")


if __name__ == "__main__":
    main()
